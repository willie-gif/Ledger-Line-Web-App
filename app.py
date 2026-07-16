"""
Ledger Line — a budgeting and investing dashboard, connected together.
Run with:  streamlit run app.py
Data is saved to ledger_line_data.json next to this file, so it persists
between runs.
"""

import csv
import io
import json
import os
import re
import uuid
from datetime import date, datetime

import streamlit as st

try:
    import requests
except ImportError:
    requests = None

try:
    import plotly.graph_objects as go
except ImportError:
    go = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import firebase_admin
    from firebase_admin import credentials as fb_credentials, firestore as fb_firestore
except ImportError:
    firebase_admin = None
    fb_credentials = None
    fb_firestore = None

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "ledger_line_data.json")
FIREBASE_KEY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "firebase-key.json")
FIRESTORE_COLLECTION = "ledger_line"
FEEDBACK_COLLECTION = "ledger_line_feedback"
FIRESTORE_DOCUMENT = "data"


def get_or_create_anon_id():
    """Gives each anonymous visitor a private, isolated data record.

    The id lives in the page's URL as a query param (?uid=...) rather than a
    login — nothing personal is collected. Returns (id, is_new_visitor).
    """
    existing = st.query_params.get("uid")
    if existing:
        return existing, False
    new_id = uuid.uuid4().hex[:12]
    st.query_params["uid"] = new_id
    return new_id, True


def _firestore_doc_id(anon_id):
    return anon_id or FIRESTORE_DOCUMENT


def _local_data_path(anon_id):
    if anon_id:
        return os.path.join(os.path.dirname(DATA_FILE), f"ledger_line_data_{anon_id}.json")
    return DATA_FILE


@st.cache_resource(show_spinner=False)
def get_firestore_client():
    """Returns a Firestore client if Firebase is configured, otherwise None.

    Looks for credentials in this order:
    1. Streamlit secrets under a [firebase] table (recommended for deployed apps).
    2. A local firebase-key.json service account file next to this script.
    If neither is present, returns None so the app falls back to local file storage.
    Any error along the way is stashed in st.session_state['firestore_error'] so
    the sidebar diagnostics can show the real reason instead of a plain ❌.
    """
    if firebase_admin is None:
        _set_firestore_error("firebase-admin package isn't installed (pip install firebase-admin)")
        return None
    try:
        if not firebase_admin._apps:
            cred = None
            secrets_error = None
            try:
                if "firebase" in st.secrets:
                    cred = fb_credentials.Certificate(dict(st.secrets["firebase"]))
            except Exception as e:
                secrets_error = str(e)
                cred = None
            if cred is None and os.path.exists(FIREBASE_KEY_FILE):
                cred = fb_credentials.Certificate(FIREBASE_KEY_FILE)
            if cred is None:
                if secrets_error:
                    _set_firestore_error(f"Found a [firebase] section in secrets.toml but couldn't use it: {secrets_error}")
                else:
                    _set_firestore_error(
                        "No credentials found — no [firebase] section in secrets.toml and no firebase-key.json file next to app.py"
                    )
                return None
            firebase_admin.initialize_app(cred)
        client = fb_firestore.client()
        _set_firestore_error(None)
        return client
    except Exception as e:
        _set_firestore_error(f"{type(e).__name__}: {e}")
        return None


_FIRESTORE_ERROR = None


def _set_firestore_error(msg):
    global _FIRESTORE_ERROR
    _FIRESTORE_ERROR = msg


def get_firebase_web_api_key():
    """The 'Web API Key' from Firebase Project Settings — different from the
    service account JSON. Needed because the Admin SDK deliberately can't
    verify passwords itself; only the REST Identity Toolkit API or a client-side
    SDK can. Looked up from st.secrets first, then an environment variable."""
    try:
        if "firebase_web_api_key" in st.secrets:
            return st.secrets["firebase_web_api_key"]
    except Exception:
        pass
    return os.environ.get("FIREBASE_WEB_API_KEY")


def get_admin_email():
    """The one account allowed to see the aggregate usage overview. Configure
    via secrets.toml (admin_email = "you@example.com") or an environment
    variable — there's no default, so the Admin page stays hidden until set."""
    try:
        if "admin_email" in st.secrets:
            return st.secrets["admin_email"]
    except Exception:
        pass
    return os.environ.get("LEDGER_LINE_ADMIN_EMAIL")


def auth_available():
    return requests is not None and get_firebase_web_api_key() is not None and get_firestore_client() is not None


def auth_diagnostics():
    """Returns a list of (label, ok) pairs so setup problems are visible instead
    of silently falling back to anonymous mode with no explanation."""
    firestore_ok = get_firestore_client() is not None
    firestore_label = "Firestore connected (service account + Firestore enabled)"
    if not firestore_ok and _FIRESTORE_ERROR:
        firestore_label += f" — {_FIRESTORE_ERROR}"
    return [
        ("requests library installed", requests is not None),
        ("Web API key found in secrets.toml", get_firebase_web_api_key() is not None),
        (firestore_label, firestore_ok),
    ]


EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

_AUTH_ERROR_MESSAGES = {
    "EMAIL_EXISTS": "An account with that email already exists — try logging in instead.",
    "EMAIL_NOT_FOUND": "No account found with that email.",
    "INVALID_PASSWORD": "Incorrect password.",
    "INVALID_LOGIN_CREDENTIALS": "Incorrect email or password.",
    "USER_DISABLED": "This account has been disabled.",
    "WEAK_PASSWORD": "Password should be at least 6 characters.",
    "INVALID_EMAIL": "That doesn't look like a valid email address.",
    "TOO_MANY_ATTEMPTS_TRY_LATER": "Too many attempts — please wait a bit and try again.",
}


def _friendly_auth_error(response_json):
    code = response_json.get("error", {}).get("message", "")
    for key, msg in _AUTH_ERROR_MESSAGES.items():
        if code.startswith(key):
            return msg
    return "Something went wrong. Please try again."


def firebase_sign_up(email, password):
    """Creates a new email/password account via the REST Identity Toolkit API.
    Returns (success, id_token_or_error_message)."""
    url = f"https://identitytoolkit.googleapis.com/v1/accounts:signUp?key={get_firebase_web_api_key()}"
    try:
        resp = requests.post(url, json={"email": email, "password": password, "returnSecureToken": True}, timeout=10)
        data = resp.json()
        if resp.ok:
            return True, data
        return False, _friendly_auth_error(data)
    except requests.RequestException:
        return False, "Couldn't reach Firebase — check your internet connection and try again."


def firebase_sign_in(email, password):
    """Logs in with email/password via the REST Identity Toolkit API.
    Returns (success, id_token_data_or_error_message)."""
    url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={get_firebase_web_api_key()}"
    try:
        resp = requests.post(url, json={"email": email, "password": password, "returnSecureToken": True}, timeout=10)
        data = resp.json()
        if resp.ok:
            return True, data
        return False, _friendly_auth_error(data)
    except requests.RequestException:
        return False, "Couldn't reach Firebase — check your internet connection and try again."


def firebase_refresh_token(refresh_token):
    """Exchanges a refresh token for a new ID token, so a visitor doesn't have to
    log in again every time their ~1 hour ID token expires. Note this endpoint's
    response uses different field names (access_token/refresh_token) than the
    sign-in/sign-up endpoints (idToken/refreshToken) — a real Firebase quirk.
    Returns (success, data_or_error_message)."""
    url = f"https://securetoken.googleapis.com/v1/token?key={get_firebase_web_api_key()}"
    try:
        resp = requests.post(url, data={"grant_type": "refresh_token", "refresh_token": refresh_token}, timeout=10)
        data = resp.json()
        if resp.ok:
            return True, data
        return False, _friendly_auth_error(data)
    except requests.RequestException:
        return False, "Couldn't reach Firebase — check your internet connection and try again."


def firebase_send_password_reset(email):
    """Asks Firebase to email the user a password reset link — Firebase sends
    and hosts the actual reset page itself, no email service needed on our end.
    Returns (success, data_or_error_message)."""
    url = f"https://identitytoolkit.googleapis.com/v1/accounts:sendOobCode?key={get_firebase_web_api_key()}"
    try:
        resp = requests.post(url, json={"requestType": "PASSWORD_RESET", "email": email}, timeout=10)
        data = resp.json()
        if resp.ok:
            return True, data
        return False, _friendly_auth_error(data)
    except requests.RequestException:
        return False, "Couldn't reach Firebase — check your internet connection and try again."


def verify_firebase_token(id_token):
    """Cryptographically verifies a Firebase ID token server-side (via the Admin
    SDK, checked against Google's public keys) and returns the verified uid, or
    None if the token is missing, invalid, or expired. This is what makes it safe
    to carry the token in the URL — unlike a raw uid, it can't be forged, and it
    expires (~1 hour) rather than granting indefinite access."""
    if firebase_admin is None:
        return None
    get_firestore_client()  # ensures the Firebase app is initialized
    try:
        from firebase_admin import auth as fb_auth
        decoded = fb_auth.verify_id_token(id_token)
        return decoded.get("uid")
    except Exception:
        return None

TYPE_LABELS = {
    "stock": "Stocks",
    "etf": "ETFs / funds",
    "bond": "Bonds",
    "crypto": "Crypto",
    "cash": "Cash",
    "other": "Other",
}

TYPE_COLORS = {
    "stock": "#2563EB",
    "etf": "#16A34A",
    "bond": "#7C3AED",
    "crypto": "#DC2626",
    "cash": "#F59E0B",
    "other": "#0891B2",
}

BROKERS = [
    # United States
    {"region": "United States", "name": "Fidelity", "note": "Full-service broker, no account minimum, $0 commissions on US stocks and ETFs.", "url": "https://www.fidelity.com"},
    {"region": "United States", "name": "Charles Schwab", "note": "Full-service broker with $0 commissions and extensive branch support.", "url": "https://www.schwab.com"},
    {"region": "United States", "name": "Vanguard", "note": "Known for low-cost index funds and ETFs, popular for long-term investing.", "url": "https://investor.vanguard.com"},
    {"region": "United States", "name": "Interactive Brokers", "note": "Broad market access and low margin rates, geared toward active traders.", "url": "https://www.interactivebrokers.com"},
    {"region": "United States", "name": "Robinhood", "note": "Mobile-first app with commission-free stocks, ETFs, options and crypto.", "url": "https://robinhood.com"},
    {"region": "United States", "name": "E*TRADE", "note": "Commission-free trading with a robo-advisor option, part of Morgan Stanley.", "url": "https://us.etrade.com"},
    {"region": "United States", "name": "Ally Invest", "note": "Self-directed or robo-advisor investing, integrates with Ally Bank.", "url": "https://www.ally.com/invest/"},
    {"region": "United States", "name": "SoFi Invest", "note": "Beginner-friendly app with no account minimums.", "url": "https://www.sofi.com/invest/"},
    {"region": "United States", "name": "M1 Finance", "note": "Automated, customizable portfolios (\"pies\") with no trading commissions.", "url": "https://m1.com"},
    # Kenya — NSE stockbrokers (full official roster of NSE trading participants)
    {"region": "Kenya — NSE stockbrokers", "name": "Dyer & Blair Investment Bank", "note": "One of Kenya's oldest investment banks, offering stockbroking, advisory and asset management.", "url": "https://www.dyerandblair.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Francis Drummond & Company", "note": "One of the original NSE founding member stockbroking firms.", "url": "https://www.drummond.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "Suntra Investment Bank", "note": "CMA-licensed stockbroker and NSE member offering stocks, bonds and advisory.", "url": "https://suntra.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "Omni Marche Securite (OMS) Africa", "note": "NSE trading participant based in Upper Hill, Nairobi.", "url": "https://www.omsafrica.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "SBG Securities", "note": "Stockbroking arm of Stanbic Bank Kenya.", "url": "https://www.sbgsecurities.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "Kingdom Securities", "note": "Stockbrokerage arm of Co-operative Bank of Kenya.", "url": "https://kingdomsecurities.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "AIB-AXYS Africa", "note": "NSE stockbroker with regional East African reach.", "url": "https://www.aib-axysafrica.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "ABC Capital", "note": "Stockbrokerage arm of ABC Bank, an NSE member since 1954.", "url": "https://www.abccapital.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "Sterling Capital", "note": "NSE stockbroker with an online and mobile trading platform.", "url": "https://www.sterlingib.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Pergamon Investment Bank", "note": "NSE trading participant based in Lavington, Nairobi.", "url": "https://pib.africa"},
    {"region": "Kenya — NSE stockbrokers", "name": "Faida Investment Bank", "note": "NSE stockbroker and investment bank with a mobile trading app.", "url": "https://www.fib.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "Standard Investment Bank (SIB)", "note": "CMA-licensed NSE stockbroker and investment bank; also runs money market and Shariah funds.", "url": "https://sib.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "Kestrel Capital", "note": "One of the longest-running NSE stockbrokers, known for equity research.", "url": "https://www.kestrelcapital.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Renaissance Capital (Kenya)", "note": "Kenyan arm of the pan-emerging-markets investment bank Renaissance Capital.", "url": "https://www.rencap.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Genghis Capital", "note": "NSE stockbroker and investment bank.", "url": "https://www.genghis-capital.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "NCBA Investment Bank", "note": "Investment banking arm of NCBA Group.", "url": "https://investment-bank.ncbagroup.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Equity Investment Bank", "note": "Investment banking arm of Equity Group.", "url": "https://www.equitybankgroup.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "KCB Investment Bank", "note": "Stockbrokerage and investment banking arm of KCB Group.", "url": "https://ke.kcbgroup.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Absa Securities", "note": "Investment banking arm of Absa Bank Kenya, part of the pan-African Absa Group.", "url": "https://www.absabank.co.ke/absa-securities-limited/"},
    {"region": "Kenya — NSE stockbrokers", "name": "Capital A Investment Bank", "note": "NSE trading participant based in Parklands, Nairobi.", "url": "https://www.caib.co.ke"},
    {"region": "Kenya — NSE stockbrokers", "name": "EFG Hermes Kenya", "note": "Kenyan arm of the pan-regional investment bank EFG Hermes.", "url": "https://www.efg-hermes.com"},
    {"region": "Kenya — NSE stockbrokers", "name": "Dry Associates", "note": "Independent investment advisory and stockbrokerage firm.", "url": "https://www.dryassociates.com"},
    # Kenya — fund managers & money market funds
    {"region": "Kenya — fund managers & money market funds", "name": "Cytonn Asset Managers", "note": "Fund manager known for high-yielding money market and fixed income funds.", "url": "https://cytonn.com"},
    {"region": "Kenya — fund managers & money market funds", "name": "CIC Asset Management", "note": "Subsidiary of CIC Insurance Group; one of the largest money market funds by assets under management.", "url": "https://ke.cicinsurancegroup.com"},
    {"region": "Kenya — fund managers & money market funds", "name": "Britam Asset Managers", "note": "Asset management arm of Britam Holdings, offering unit trusts and money market funds.", "url": "https://ke.britam.com"},
    {"region": "Kenya — fund managers & money market funds", "name": "Old Mutual Investment Group", "note": "Runs the Old Mutual and Zimele money market and unit trust funds.", "url": "https://www.oldmutual.co.ke"},
    {"region": "Kenya — fund managers & money market funds", "name": "ICEA Lion Asset Management", "note": "One of Kenya's oldest fund managers, offering money market and other unit trust funds.", "url": "https://icealion.com"},
    {"region": "Kenya — fund managers & money market funds", "name": "Sanlam Allianz Investments", "note": "Fund management arm of Sanlam Kenya, following its 2025 merger with Allianz.", "url": "https://www.sanlamallianzinvestments.com"},
    {"region": "Kenya — fund managers & money market funds", "name": "Madison Investment Managers", "note": "Manages the Madison Money Market Fund and other unit trusts.", "url": "https://www.madison.co.ke"},
    {"region": "Kenya — fund managers & money market funds", "name": "Zimele Asset Management", "note": "One of Kenya's oldest unit trust managers, known for a KES 100 minimum investment.", "url": "https://www.zimele.co.ke"},
    {"region": "Kenya — fund managers & money market funds", "name": "GenAfrica Asset Managers", "note": "Established fund manager (formerly Genesis Kenya) focused on institutional and retail portfolios.", "url": "https://genafrica.com"},
    {"region": "Kenya — fund managers & money market funds", "name": "Nabo Capital", "note": "Investment manager now majority-owned by Rock Investment Bank following a 2026 stake sale by Centum.", "url": "https://www.nabocapital.com"},
    # Kenya — mobile investing apps
    {"region": "Kenya — mobile investing apps", "name": "Ndovu", "note": "CMA-licensed app for investing in ETFs, including US-listed funds, from a phone.", "url": "https://ndovu.co"},
    {"region": "Kenya — mobile investing apps", "name": "Etica", "note": "Low-minimum money market fund app popular with first-time investors.", "url": "https://etica.co.ke"},
    {"region": "Kenya — mobile investing apps", "name": "Chumz", "note": "Savings and investing app that rounds up spending into a money market fund.", "url": "https://chumz.io"},
]

BROKER_REGIONS = list(dict.fromkeys(b["region"] for b in BROKERS))  # preserves order, dedups

RISK_QUESTIONS = [
    {
        "question": "When do you expect to need this invested money?",
        "options": [
            ("Within 1 year", 1),
            ("1–3 years", 2),
            ("3–7 years", 3),
            ("7+ years", 4),
        ],
    },
    {
        "question": "If your investments dropped 20% in a few months, what would you most likely do?",
        "options": [
            ("Sell everything to stop further loss", 1),
            ("Sell some to reduce risk", 2),
            ("Hold and wait it out", 3),
            ("Buy more while prices are lower", 4),
        ],
    },
    {
        "question": "How would you describe your investing experience?",
        "options": [
            ("None — this would be my first investment", 1),
            ("Some — I have savings or a money market account", 2),
            ("Moderate — I've invested in stocks or funds before", 3),
            ("Experienced — I actively manage a diversified portfolio", 4),
        ],
    },
    {
        "question": "Do you have an emergency fund covering at least 3 months of expenses, separate from this money?",
        "options": [
            ("No emergency fund yet", 1),
            ("Partial — 1–2 months", 2),
            ("Yes — 3–6 months", 3),
            ("Yes — 6+ months", 4),
        ],
    },
    {
        "question": "What's your primary goal for this money?",
        "options": [
            ("Preserve what I have, avoid any loss", 1),
            ("Steady, modest growth", 2),
            ("Grow it significantly — some ups and downs are fine", 3),
            ("Maximize growth — comfortable with big swings", 4),
        ],
    },
    {
        "question": "What portion of your total savings or net worth would this investment represent?",
        "options": [
            ("Most or all of it", 1),
            ("A large portion", 2),
            ("A moderate portion", 3),
            ("A small portion — I have plenty elsewhere", 4),
        ],
    },
]

RISK_ALLOCATIONS = {
    "Conservative": {"cash": 60, "bond": 25, "etf": 10, "stock": 5, "crypto": 0},
    "Moderate": {"cash": 25, "bond": 20, "etf": 35, "stock": 15, "crypto": 5},
    "Aggressive": {"cash": 5, "bond": 10, "etf": 35, "stock": 40, "crypto": 10},
}

RISK_LEVEL_TABS = {
    "Conservative": ["Kenya — fund managers & money market funds", "Kenya — mobile investing apps"],
    "Moderate": ["Kenya — fund managers & money market funds", "Kenya — NSE stockbrokers"],
    "Aggressive": ["Kenya — NSE stockbrokers", "United States"],
}

RISK_LEVEL_BLURB = {
    "Conservative": "You're prioritizing protecting what you have over chasing growth — that's a completely reasonable stance, especially with a shorter time horizon or less experience.",
    "Moderate": "You're comfortable with some ups and downs in exchange for better long-term growth, without going all-in on higher-risk assets.",
    "Aggressive": "You're comfortable with significant short-term swings in pursuit of higher long-term growth — typically a better fit with a longer time horizon and money you won't need soon.",
}


def score_risk_answers(points):
    """points is a list of the point values chosen for each RISK_QUESTIONS item.
    Returns (score, level). Range is 6-24 across 6 questions, split into thirds."""
    score = sum(points)
    if score <= 11:
        level = "Conservative"
    elif score <= 17:
        level = "Moderate"
    else:
        level = "Aggressive"
    return score, level


PALETTE_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Fraunces:opsz,wght@9..144,500;9..144,600&family=Inter:wght@400;500;600&family=IBM+Plex+Mono:wght@400;500&display=swap');

:root {
  --paper: #0B1120;
  --card: #182338;
  --ink: #F1F5F9;
  --ink-soft: #94A3B8;
  --brass: #3B82F6;
  --forest: #22C55E;
  --brick: #EF4444;
  --purple: #A78BFA;
  --line: #2E3B52;
}
html, body { background: #0B1120; }
header[data-testid="stHeader"] { background: transparent; }
.stApp {
  background: linear-gradient(160deg, #0B1120 0%, #131C33 45%, #1B1436 100%);
  background-attachment: fixed;
}
section[data-testid="stSidebar"] {
  background: linear-gradient(180deg, #0B1120 0%, #131C33 55%, #1B1436 100%) !important;
  border-right: 1px solid var(--line);
}
section[data-testid="stSidebar"] * { color: var(--ink) !important; }
section[data-testid="stSidebar"] .stCaption, section[data-testid="stSidebar"] small { color: var(--ink-soft) !important; }
section[data-testid="stSidebar"] button {
  background: var(--card) !important; border: 1px solid var(--line) !important; color: var(--ink) !important;
}
section[data-testid="stSidebar"] button:hover { border-color: var(--brass) !important; }
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzone"] {
  background: var(--card) !important; border: 1px dashed var(--line) !important;
}
section[data-testid="stSidebar"] [data-testid="stFileUploaderDropzoneInstructions"] * { color: var(--ink-soft) !important; }
section[data-testid="stSidebar"] [data-testid="stBaseButton-secondary"] { background: var(--card) !important; }
section[data-testid="stSidebar"] div[data-baseweb="radio"] { background: transparent !important; }
section[data-testid="stSidebar"] hr { border-color: var(--line) !important; }
section[data-testid="stSidebar"] [data-testid="stAlert"] { background: var(--card) !important; border: 1px solid var(--line) !important; }
h1, h2, h3 { font-family: 'Fraunces', serif !important; color: var(--ink); font-weight: 700 !important; }
h4 { color: var(--brass) !important; font-weight: 700 !important; }
body, p, div, span, label { font-family: 'Inter', sans-serif; color: var(--ink); }
.ll-mono { font-family: 'IBM Plex Mono', monospace; }
.ll-card {
  background: var(--card); border: 1px solid var(--line); border-left: 5px solid var(--brass);
  border-radius: 4px; padding: 16px 18px; margin-bottom: 14px;
  box-shadow: 0 2px 10px rgba(0,0,0,0.35);
}
.ll-card.blue { border-left-color: var(--brass); }
.ll-card.red { border-left-color: var(--brick); }
.ll-card.green { border-left-color: var(--forest); }
.ll-card.purple { border-left-color: var(--purple); }
.ll-stat-label { font-size: 11px; text-transform: uppercase; letter-spacing: .06em; color: var(--ink-soft); font-weight: 600; }
.ll-stat-val { font-family: 'IBM Plex Mono', monospace; font-size: 24px; font-weight: 700; color: var(--ink); }
.ll-row { display:flex; justify-content:space-between; align-items:center; padding:8px 0; border-bottom:1px solid var(--line); font-size:14px; }
.ll-meta { color: var(--ink-soft); font-size: 12px; }
.ll-tag { background: var(--brass); color: #FFFFFF; border-radius: 3px; padding: 2px 9px; font-size: 11px; font-weight: 600; }

/* --- Main page widgets --- */
.stTextInput input, .stNumberInput input, .stDateInput input, textarea,
div[data-baseweb="input"] > div, div[data-baseweb="select"] > div, div[data-baseweb="base-input"] {
  background: var(--card) !important; border: 1px solid var(--line) !important; color: var(--ink) !important;
}
div[data-baseweb="popover"], div[data-baseweb="popover"] > div,
div[data-baseweb="popover"] ul, div[data-baseweb="menu"],
ul[role="listbox"], div[role="listbox"] {
  background: var(--card) !important; color: var(--ink) !important; border: 1px solid var(--line) !important;
}
div[data-baseweb="popover"] li, li[role="option"], div[role="option"] {
  background: var(--card) !important; color: var(--ink) !important;
}
div[data-baseweb="popover"] li:hover, li[role="option"]:hover, div[role="option"]:hover {
  background: var(--paper) !important;
}
.stButton button, .stFormSubmitButton button, .stDownloadButton button, .stLinkButton a,
[data-testid="stBaseButton-secondary"] {
  background: var(--card) !important; border: 1px solid var(--line) !important; color: var(--ink) !important;
}
.stButton button:hover, .stFormSubmitButton button:hover, .stDownloadButton button:hover, .stLinkButton a:hover,
[data-testid="stBaseButton-secondary"]:hover {
  border-color: var(--brass) !important; color: var(--brass) !important;
}
button[kind="primary"], button[data-testid="stBaseButton-primary"] {
  background: var(--brass) !important; border: 1px solid var(--brass) !important; color: #FFFFFF !important;
}
[data-testid="stFileUploaderDropzone"] { background: var(--card) !important; border: 1px dashed var(--line) !important; }
[data-testid="stFileUploaderDropzoneInstructions"] * { color: var(--ink-soft) !important; }
[data-testid="stAlert"] { background: var(--card) !important; border: 1px solid var(--line) !important; }
[data-testid="stNumberInputStepUp"], [data-testid="stNumberInputStepDown"] {
  background: var(--card) !important; border-color: var(--line) !important; color: var(--ink) !important;
}
div[data-baseweb="tab-list"] { background: transparent !important; border-bottom: 1px solid var(--line) !important; }
button[data-baseweb="tab"] { color: var(--ink-soft) !important; }
button[data-baseweb="tab"][aria-selected="true"] { color: var(--brass) !important; }
div[data-baseweb="tab-highlight"] { background-color: var(--brass) !important; }
[data-testid="stProgress"] > div > div { background: var(--line) !important; }
div[data-baseweb="radio"] label { color: var(--ink) !important; }
hr { border-color: var(--line) !important; }
a { color: var(--brass) !important; }
[data-testid="stDataFrame"], [data-testid="stElementToolbar"] { background: var(--card) !important; }
[data-testid="stDataFrameResizable"] { border: 1px solid var(--line) !important; border-radius: 4px; }

/* --- Modal dialog (reset confirmation) --- */
div[data-testid="stDialog"] div[role="dialog"] {
  background: var(--card) !important; border: 1px solid var(--line) !important;
  box-shadow: 0 8px 30px rgba(0,0,0,0.5) !important;
}
div[data-testid="stDialog"] div[role="dialog"] * { color: var(--ink) !important; }
div[data-testid="stDialog"] [data-testid="stAlert"] { background: var(--paper) !important; border: 1px solid var(--line) !important; }
</style>
"""

# ---------------------------------------------------------------------------
# Data persistence
# ---------------------------------------------------------------------------

def uid():
    return uuid.uuid4().hex[:8]


def paginated(items, session_key, page_size=10):
    """Shows only the first N items with a 'Show more' button to reveal more,
    instead of rendering a potentially long list all at once every rerun.
    Call render_show_more(...) right after displaying the returned slice."""
    if session_key not in st.session_state:
        st.session_state[session_key] = page_size
    limit = st.session_state[session_key]
    return items[:limit], max(0, len(items) - limit)


def render_show_more(session_key, remaining, page_size=10):
    if remaining > 0:
        if st.button(f"Show {min(remaining, page_size)} more ({remaining} remaining)", key=f"more_{session_key}"):
            st.session_state[session_key] += page_size
            st.rerun()


def default_data():
    return {
        "income": 5200,
        "categories": [
            {"id": uid(), "name": "Rent", "budgeted": 1600},
            {"id": uid(), "name": "Groceries", "budgeted": 500},
            {"id": uid(), "name": "Transport", "budgeted": 200},
            {"id": uid(), "name": "Dining out", "budgeted": 250},
            {"id": uid(), "name": "Utilities", "budgeted": 180},
        ],
        "transactions": [],
        "holdings": [
            {"id": uid(), "name": "S&P 500 index fund", "type": "etf", "value": 14200},
            {"id": uid(), "name": "Emergency cash", "type": "cash", "value": 6000},
            {"id": uid(), "name": "Individual stocks", "type": "stock", "value": 3100},
        ],
        "contributions": [],
        "net_worth_goal": 100000,
        "goals": [
            {"id": uid(), "name": "Emergency fund", "target": 15000, "saved": 4200, "notified": False},
        ],
        "recurring_transactions": [],
        "net_worth_history": [],
        "usage": {"session_count": 0, "last_active": None, "opened_broker_directory": False},
        "risk_profile": None,
        "feedback_history": [],
        "feedback_prompt_count": 0,
    }


def _clean_data(data):
    data.setdefault("goals", [])
    for g in data["goals"]:
        g.setdefault("notified", False)
    data.setdefault("net_worth_goal", 0)
    data.setdefault("recurring_transactions", [])
    data.setdefault("net_worth_history", [])
    data.setdefault("usage", {"session_count": 0, "last_active": None, "opened_broker_directory": False})
    data.setdefault("risk_profile", None)
    data.setdefault("feedback_history", [])
    data.setdefault("feedback_prompt_count", 0)
    return data


def load_data(anon_id=None):
    db = get_firestore_client()
    if db is not None:
        try:
            snap = db.collection(FIRESTORE_COLLECTION).document(_firestore_doc_id(anon_id)).get()
            if snap.exists:
                return _clean_data(snap.to_dict())
            fresh = default_data()
            save_data(fresh, anon_id)
            return fresh
        except Exception:
            pass  # fall through to local file if Firestore read fails

    path = _local_data_path(anon_id)
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                return _clean_data(json.load(f))
        except Exception:
            pass
    return default_data()


def save_data(data, anon_id=None):
    db = get_firestore_client()
    if db is not None:
        try:
            db.collection(FIRESTORE_COLLECTION).document(_firestore_doc_id(anon_id)).set(data)
            return
        except Exception:
            pass  # fall through to local file if Firestore write fails

    with open(_local_data_path(anon_id), "w") as f:
        json.dump(data, f, indent=2)


def load_current_data():
    """Session-aware wrapper used by the running app; reads the visitor's own id
    from st.session_state, which Streamlit correctly isolates per browser session
    (unlike a plain module-level variable, which would be unsafe with concurrent users)."""
    return load_data(st.session_state.get("anon_id"))


def save_current_data(data):
    save_data(data, st.session_state.get("anon_id"))


def using_firestore():
    return get_firestore_client() is not None


def submit_feedback(data, rating, comment):
    """Always saves the rating into the user's own data (so it's never lost,
    even in local-file-only mode), and additionally mirrors it into a shared
    Firestore collection — separate from each user's private document — so
    every rating can be reviewed together in the admin page."""
    entry = {"id": uid(), "rating": rating, "comment": comment, "date": str(date.today())}
    data.setdefault("feedback_history", []).append(entry)
    save_current_data(data)

    db = get_firestore_client()
    if db is not None:
        try:
            payload = dict(entry)
            payload["user_id"] = st.session_state.get("anon_id")
            payload["email"] = st.session_state.get("user_email")
            db.collection(FEEDBACK_COLLECTION).document(f"{payload['user_id'] or 'anon'}_{entry['id']}").set(payload)
        except Exception:
            pass  # local copy above already saved either way


if hasattr(st, "dialog"):
    @st.dialog("Reset all data?")
    def reset_confirmation_dialog():
        st.write("This will permanently erase your income, categories, transactions, holdings, contributions, and goals.")
        st.write("**This can't be undone.** Consider exporting a backup first if you're not sure.")
        col1, col2 = st.columns(2)
        if col1.button("Cancel", use_container_width=True):
            st.session_state.show_reset_dialog = False
            st.rerun()
        if col2.button("Yes, reset everything", type="primary", use_container_width=True):
            st.session_state.data = default_data()
            save_current_data(st.session_state.data)
            st.session_state.show_reset_dialog = False
            st.rerun()
else:
    def reset_confirmation_dialog():
        st.sidebar.warning("This will permanently erase all data. This can't be undone.")
        col1, col2 = st.sidebar.columns(2)
        if col1.button("Cancel"):
            st.session_state.show_reset_dialog = False
            st.rerun()
        if col2.button("Yes, reset everything"):
            st.session_state.data = default_data()
            save_current_data(st.session_state.data)
            st.session_state.show_reset_dialog = False
            st.rerun()


if hasattr(st, "dialog"):
    @st.dialog("Rate Ledger Line")
    def rating_dialog(data):
        st.write("How's the app working for you so far?")
        if hasattr(st, "feedback"):
            selected = st.feedback("stars", key="rating_stars")
            rating = (selected + 1) if selected is not None else None
        else:
            rating = st.radio("Rating (1 = poor, 5 = great)", [1, 2, 3, 4, 5], horizontal=True, index=None, key="rating_radio_fallback")
        comment = st.text_area("Anything you'd like to add? (optional)", key="rating_comment")
        col1, col2 = st.columns(2)
        if col1.button("Maybe later", use_container_width=True):
            st.session_state.show_rating_dialog = False
            st.rerun()
        if col2.button("Submit", type="primary", use_container_width=True):
            if rating is None:
                st.error("Please pick a rating first.")
            else:
                submit_feedback(data, rating, comment)
                st.session_state.show_rating_dialog = False
                st.success("Thanks for the feedback!")
                st.rerun()
else:
    def rating_dialog(data):
        st.sidebar.markdown("**Rate Ledger Line**")
        rating = st.sidebar.radio("Rating (1 = poor, 5 = great)", [1, 2, 3, 4, 5], index=None, key="rating_radio_fallback")
        comment = st.sidebar.text_area("Anything you'd like to add? (optional)", key="rating_comment")
        col1, col2 = st.sidebar.columns(2)
        if col1.button("Maybe later"):
            st.session_state.show_rating_dialog = False
            st.rerun()
        if col2.button("Submit"):
            if rating is None:
                st.sidebar.error("Please pick a rating first.")
            else:
                submit_feedback(data, rating, comment)
                st.session_state.show_rating_dialog = False
                st.sidebar.success("Thanks for the feedback!")
                st.rerun()


def fmt(n):
    sign = "-" if n < 0 else ""
    return f"{sign}${abs(round(n)):,}"


CSV_COLUMNS = ["type", "id", "name", "date", "amount", "category_id", "holding_id",
               "note", "budgeted", "value", "holding_type", "target", "saved",
               "notified", "income", "net_worth_goal"]


def _num(v, default=0):
    try:
        if v is None or v == "":
            return default
        return float(v)
    except (ValueError, TypeError):
        return default


def data_to_csv(data):
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=CSV_COLUMNS, restval="")
    writer.writeheader()
    writer.writerow({"type": "setting", "income": data["income"], "net_worth_goal": data["net_worth_goal"]})
    for c in data["categories"]:
        writer.writerow({"type": "category", "id": c["id"], "name": c["name"], "budgeted": c["budgeted"]})
    for t in data["transactions"]:
        writer.writerow({"type": "transaction", "id": t["id"], "date": t["date"],
                          "category_id": t["categoryId"], "amount": t["amount"], "note": t.get("note", "")})
    for h in data["holdings"]:
        writer.writerow({"type": "holding", "id": h["id"], "name": h["name"],
                          "holding_type": h["type"], "value": h["value"]})
    for c in data["contributions"]:
        writer.writerow({"type": "contribution", "id": c["id"], "date": c["date"], "amount": c["amount"],
                          "holding_id": c.get("holdingId") or "", "note": c.get("note", "")})
    for g in data["goals"]:
        writer.writerow({"type": "goal", "id": g["id"], "name": g["name"], "target": g["target"],
                          "saved": g["saved"], "notified": g.get("notified", False)})
    return buf.getvalue()


def csv_to_data(text):
    result = {"income": 0, "categories": [], "transactions": [], "holdings": [],
              "contributions": [], "net_worth_goal": 0, "goals": []}
    reader = csv.DictReader(io.StringIO(text))
    for row in reader:
        kind = row.get("type", "")
        if kind == "setting":
            result["income"] = _num(row.get("income"))
            result["net_worth_goal"] = _num(row.get("net_worth_goal"))
        elif kind == "category":
            result["categories"].append({"id": row["id"], "name": row["name"], "budgeted": _num(row.get("budgeted"))})
        elif kind == "transaction":
            result["transactions"].append({
                "id": row["id"], "date": row.get("date", ""), "categoryId": row.get("category_id", ""),
                "amount": _num(row.get("amount")), "note": row.get("note", ""),
            })
        elif kind == "holding":
            result["holdings"].append({
                "id": row["id"], "name": row["name"], "type": row.get("holding_type", "other"),
                "value": _num(row.get("value")),
            })
        elif kind == "contribution":
            result["contributions"].append({
                "id": row["id"], "date": row.get("date", ""), "amount": _num(row.get("amount")),
                "note": row.get("note", ""), "holdingId": row.get("holding_id") or None,
            })
        elif kind == "goal":
            result["goals"].append({
                "id": row["id"], "name": row["name"], "target": _num(row.get("target")),
                "saved": _num(row.get("saved")), "notified": str(row.get("notified")).strip().lower() == "true",
            })
    return result


# ---------------------------------------------------------------------------
# Derived calculations
# ---------------------------------------------------------------------------

def current_month_str():
    return date.today().strftime("%Y-%m")


def month_key(date_str):
    return date_str[:7] if date_str else ""


def transactions_for_month(data, month=None):
    """month=None returns every transaction ever logged (the original, all-time
    behavior). Pass an explicit 'YYYY-MM' string to scope to just that month."""
    if month is None:
        return data["transactions"]
    return [t for t in data["transactions"] if month_key(t["date"]) == month]


def available_transaction_months(data):
    months = {month_key(t["date"]) for t in data["transactions"] if t["date"]}
    months.add(current_month_str())
    return sorted(months, reverse=True)


def total_spent(data, month=None):
    return sum(t["amount"] for t in transactions_for_month(data, month))


def spent_by_category(data, cat_id, month=None):
    return sum(t["amount"] for t in transactions_for_month(data, month) if t["categoryId"] == cat_id)


def total_budgeted(data):
    return sum(c["budgeted"] for c in data["categories"])


def saved_this_month(data, month=None):
    """month=None preserves the original all-time behavior. The UI explicitly
    passes the current month so the number actually resets monthly."""
    return data["income"] - total_spent(data, month)


def total_contributed(data):
    return sum(c["amount"] for c in data["contributions"])


def generate_recurring_transactions(data):
    """Auto-logs this month's copy of each recurring transaction template, if it
    hasn't been generated yet. Returns True if anything new was created (so the
    caller only needs to save when something actually changed)."""
    month = current_month_str()
    already_generated = {
        t.get("recurringId") for t in data["transactions"]
        if month_key(t["date"]) == month and t.get("recurringId")
    }
    created_any = False
    for r in data.get("recurring_transactions", []):
        if r["id"] in already_generated:
            continue
        day = max(1, min(int(r.get("day_of_month", 1)), 28))
        tx_date = f"{month}-{day:02d}"
        data["transactions"].append({
            "id": uid(), "date": tx_date, "categoryId": r["categoryId"],
            "amount": r["amount"], "note": r.get("note", ""), "recurringId": r["id"],
        })
        created_any = True
    return created_any


def total_portfolio(data):
    return sum(h["value"] for h in data["holdings"])


def record_net_worth_snapshot(data):
    """Upserts this month's net worth snapshot. Past months stay locked in once
    the month has passed; the current month's point keeps updating in place as
    holdings change, so the chart always ends on today's real value. Returns
    True if anything actually changed (so the caller only saves when needed)."""
    month = current_month_str()
    value = total_portfolio(data)
    history = data.setdefault("net_worth_history", [])
    existing = next((h for h in history if h["month"] == month), None)
    if existing:
        if existing["value"] != value:
            existing["value"] = value
            return True
        return False
    history.append({"month": month, "value": value})
    history.sort(key=lambda h: h["month"])
    return True


def track_session_usage(data):
    """Lightweight, privacy-minimal usage signals — just enough to answer 'is
    anyone actually using this' without logging anything sensitive. Counted
    once per browser session (not every rerun) via a session_state flag."""
    changed = False
    usage = data.setdefault("usage", {"session_count": 0, "last_active": None, "opened_broker_directory": False})
    today = str(date.today())
    if not st.session_state.get("session_counted"):
        usage["session_count"] = usage.get("session_count", 0) + 1
        st.session_state.session_counted = True
        changed = True
    if usage.get("last_active") != today:
        usage["last_active"] = today
        changed = True
    return changed


def investment_suggestion(amount, data):
    """General, informational suggestions based on how much is available to invest
    this month — not personalized financial advice. Tiers are illustrative
    starting points to help someone figure out where to look, not a recommendation
    of what to actually do with their money."""
    by_type = {}
    for h in data["holdings"]:
        by_type[h["type"]] = by_type.get(h["type"], 0) + h["value"]
    has_growth_assets = any(t in by_type for t in ("stock", "etf", "bond"))
    no_holdings_yet = len(data["holdings"]) == 0

    result = None
    if amount <= 0:
        result = {
            "title": "Nothing extra to invest this month",
            "body": "Spending matched or exceeded income this month, so there isn't a surplus to put to work yet. "
                    "Worth revisiting your category budgets on the Budget page to see where there's room.",
            "tabs": [],
        }
    elif amount < 50:
        result = {
            "title": "A small but real amount",
            "body": "Some money market funds and mobile investing apps accept very low minimums — "
                    "some Kenyan platforms start around KES 100 — so even a small amount doesn't have to sit idle. "
                    "Worth a look: the **Kenya — mobile investing apps** and **fund managers & money market funds** tabs below.",
            "tabs": ["Kenya — mobile investing apps", "Kenya — fund managers & money market funds"],
        }
    elif amount < 500:
        result = {
            "title": "Enough to start building a base",
            "body": "A money market fund is a common starting point at this level — it's liquid and low-risk while "
                    "you build up further. If you're comfortable with more risk for potential growth, a small starter "
                    "position in a stock or ETF is also worth considering. Check **fund managers & money market funds** "
                    "and **Kenya — NSE stockbrokers** below.",
            "tabs": ["Kenya — fund managers & money market funds", "Kenya — NSE stockbrokers"],
        }
    elif amount < 5000:
        result = {
            "title": "Enough room to diversify a little",
            "body": "At this level, some people split between something liquid (a money market fund) and something "
                    "growth-oriented (stocks or ETFs). If you want exposure beyond the Kenyan market too, a US broker "
                    "is worth a look. Check **Kenya — NSE stockbrokers**, **fund managers & money market funds**, and "
                    "**United States** below.",
            "tabs": ["Kenya — NSE stockbrokers", "Kenya — fund managers & money market funds", "United States"],
        }
    else:
        result = {
            "title": "A meaningful amount to put to work",
            "body": "At this size, it's often worth thinking across asset classes — stocks, bonds, ETFs — and possibly "
                    "across both local (NSE) and international (US) markets, rather than one single option. This is "
                    "also a reasonable amount to run past a licensed financial advisor for guidance specific to your "
                    "situation. Check **Kenya — NSE stockbrokers**, **fund managers & money market funds**, and "
                    "**United States** below.",
            "tabs": ["Kenya — NSE stockbrokers", "Kenya — fund managers & money market funds", "United States"],
        }

    if amount > 0:
        if no_holdings_yet:
            result["body"] += " Since there aren't any holdings logged yet, this could be a good first one — add it on the Portfolio page once you've picked somewhere."
        elif not has_growth_assets:
            result["body"] += " Worth noting: everything currently logged is cash or similar — no stocks, ETFs, or bonds yet — so this could be a chance to add some growth exposure if that fits your goals."
    return result


def json_to_data(text):
    parsed = json.loads(text)
    required = ["income", "categories", "transactions", "holdings", "contributions"]
    if not all(k in parsed for k in required):
        raise ValueError("Missing required fields")
    parsed.setdefault("goals", [])
    parsed.setdefault("net_worth_goal", 0)
    for g in parsed["goals"]:
        g.setdefault("notified", False)
    return parsed


XLSX_SHEETS = {
    "Settings": ["income", "net_worth_goal"],
    "Categories": ["id", "name", "budgeted"],
    "Transactions": ["id", "date", "category_id", "amount", "note"],
    "Holdings": ["id", "name", "type", "value"],
    "Contributions": ["id", "date", "amount", "holding_id", "note"],
    "Goals": ["id", "name", "target", "saved", "notified"],
}


def data_to_xlsx(data):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Settings")
    ws.append(XLSX_SHEETS["Settings"])
    ws.append([data["income"], data["net_worth_goal"]])

    ws = wb.create_sheet("Categories")
    ws.append(XLSX_SHEETS["Categories"])
    for c in data["categories"]:
        ws.append([c["id"], c["name"], c["budgeted"]])

    ws = wb.create_sheet("Transactions")
    ws.append(XLSX_SHEETS["Transactions"])
    for t in data["transactions"]:
        ws.append([t["id"], t["date"], t["categoryId"], t["amount"], t.get("note", "")])

    ws = wb.create_sheet("Holdings")
    ws.append(XLSX_SHEETS["Holdings"])
    for h in data["holdings"]:
        ws.append([h["id"], h["name"], h["type"], h["value"]])

    ws = wb.create_sheet("Contributions")
    ws.append(XLSX_SHEETS["Contributions"])
    for c in data["contributions"]:
        ws.append([c["id"], c["date"], c["amount"], c.get("holdingId") or "", c.get("note", "")])

    ws = wb.create_sheet("Goals")
    ws.append(XLSX_SHEETS["Goals"])
    for g in data["goals"]:
        ws.append([g["id"], g["name"], g["target"], g["saved"], g.get("notified", False)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def xlsx_to_data(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    result = {"income": 0, "categories": [], "transactions": [], "holdings": [],
              "contributions": [], "net_worth_goal": 0, "goals": []}

    def rows_of(sheet_name):
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        out = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            out.append(dict(zip(header, row)))
        return out

    settings_rows = rows_of("Settings")
    if settings_rows:
        result["income"] = _num(settings_rows[0].get("income"))
        result["net_worth_goal"] = _num(settings_rows[0].get("net_worth_goal"))

    for r in rows_of("Categories"):
        result["categories"].append({"id": str(r.get("id")), "name": r.get("name") or "", "budgeted": _num(r.get("budgeted"))})
    for r in rows_of("Transactions"):
        result["transactions"].append({
            "id": str(r.get("id")), "date": str(r.get("date") or ""), "categoryId": str(r.get("category_id") or ""),
            "amount": _num(r.get("amount")), "note": r.get("note") or "",
        })
    for r in rows_of("Holdings"):
        result["holdings"].append({"id": str(r.get("id")), "name": r.get("name") or "",
                                     "type": r.get("type") or "other", "value": _num(r.get("value"))})
    for r in rows_of("Contributions"):
        hid = r.get("holding_id")
        result["contributions"].append({
            "id": str(r.get("id")), "date": str(r.get("date") or ""), "amount": _num(r.get("amount")),
            "note": r.get("note") or "", "holdingId": str(hid) if hid else None,
        })
    for r in rows_of("Goals"):
        result["goals"].append({
            "id": str(r.get("id")), "name": r.get("name") or "", "target": _num(r.get("target")),
            "saved": _num(r.get("saved")), "notified": str(r.get("notified")).strip().lower() == "true",
        })
    return result


# ---------------------------------------------------------------------------
# Page: Overview
# ---------------------------------------------------------------------------

def page_overview(data):
    st.title("Overview")
    st.caption(date.today().strftime("%B %Y"))

    this_month = current_month_str()
    spent = total_spent(data, this_month)
    saved = saved_this_month(data, this_month)
    portfolio = total_portfolio(data)
    contributed = total_contributed(data)

    cols = st.columns(4)
    stats = [
        ("Monthly income", fmt(data["income"]), "blue"),
        ("Spent", fmt(spent), "red"),
        ("Available to invest", fmt(saved), "green"),
        ("Portfolio value", fmt(portfolio), "purple"),
    ]
    for col, (label, val, accent) in zip(cols, stats):
        col.markdown(
            f'<div class="ll-card {accent}"><div class="ll-stat-label">{label}</div>'
            f'<div class="ll-stat-val">{val}</div></div>',
            unsafe_allow_html=True,
        )

    reached_goals = [g for g in data["goals"] if g["saved"] >= g["target"] and g["target"] > 0]
    if reached_goals:
        names = ", ".join(f'"{g["name"]}"' for g in reached_goals)
        st.success(f"🎉 Goal(s) fully funded: {names}")

    st.markdown("#### What could this go toward?")
    suggestion = investment_suggestion(saved, data)
    st.markdown(f"**{suggestion['title']}**")
    st.markdown(suggestion["body"])

    risk_profile = data.get("risk_profile")
    if risk_profile:
        st.caption(f"Factoring in your **{risk_profile['level']}** risk profile — view the full breakdown on the Risk Profile page.")
    else:
        if st.button("Take the 2-minute risk quiz for tailored suggestions →", key="jump_to_risk_quiz"):
            st.session_state.pending_nav = "Risk Profile"
            st.rerun()

    if suggestion["tabs"]:
        st.markdown("**Ready to start? Jump straight to a category on the Portfolio page:**")
        cols = st.columns(len(suggestion["tabs"]))
        for col, region in zip(cols, suggestion["tabs"]):
            if col.button(f"{region} →", key=f"jump_{region}", use_container_width=True):
                st.session_state.broker_region_tab = region
                st.session_state.broker_search = ""
                st.session_state.pending_nav = "Portfolio"
                st.rerun()
    st.markdown(
        "<p style='font-style: italic; color: var(--brick); font-size: 0.85rem;'>"
        "General information to help you start looking, not personalized financial advice — "
        "I'm not a financial advisor. Always compare fees and features yourself before choosing where to invest."
        "</p>",
        unsafe_allow_html=True,
    )

    st.markdown("#### Where this month's income goes")
    if go is not None:
        fig = go.Figure(go.Sankey(
            node=dict(
                label=["Income", "Spent", "Available to invest", "Contributed to portfolio"],
                color=["#F1F5F9", "#EF4444", "#22C55E", "#3B82F6"],
                pad=20, thickness=16,
            ),
            link=dict(
                source=[0, 0, 2],
                target=[1, 2, 3],
                value=[max(spent, 1), max(saved, 1), max(contributed, 1)],
                color=["rgba(155,70,50,0.35)", "rgba(63,107,76,0.35)", "rgba(168,118,62,0.35)"],
            ),
        ))
        fig.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10),
                           paper_bgcolor="rgba(0,0,0,0)", font=dict(family="Inter"))
        st.plotly_chart(fig, use_container_width=True)
        st.caption("Note: the 'contributed to portfolio' flow reflects total logged contributions, not strictly this month's — it's illustrative of how budgeting feeds investing.")
    else:
        st.info("Install plotly (`pip install plotly`) to see the income flow diagram.")

    left, right = st.columns(2)
    with left:
        st.markdown("**Recent spending**")
        recent_tx = sorted(data["transactions"], key=lambda t: t["date"], reverse=True)[:5]
        if not recent_tx:
            st.caption("No transactions logged yet. Add some on the Budget page.")
        for t in recent_tx:
            cat = next((c for c in data["categories"] if c["id"] == t["categoryId"]), None)
            name = cat["name"] if cat else "Uncategorized"
            note = f" · {t['note']}" if t.get("note") else ""
            st.markdown(
                f'<div class="ll-row"><div>{name}<div class="ll-meta">{t["date"]}{note}</div></div>'
                f'<div class="ll-mono" style="color:var(--brick);">-{fmt(t["amount"])}</div></div>',
                unsafe_allow_html=True,
            )
    with right:
        st.markdown("**Recent contributions**")
        recent_c = sorted(data["contributions"], key=lambda c: c["date"], reverse=True)[:5]
        if not recent_c:
            st.caption("No contributions logged yet. Add some on the Portfolio page.")
        for c in recent_c:
            h = next((h for h in data["holdings"] if h["id"] == c.get("holdingId")), None)
            name = h["name"] if h else "Unallocated"
            note = f" · {c['note']}" if c.get("note") else ""
            st.markdown(
                f'<div class="ll-row"><div>{name}<div class="ll-meta">{c["date"]}{note}</div></div>'
                f'<div class="ll-mono" style="color:var(--forest);">+{fmt(c["amount"])}</div></div>',
                unsafe_allow_html=True,
            )


# ---------------------------------------------------------------------------
# Page: Budget
# ---------------------------------------------------------------------------

def parse_quick_add(text, categories):
    """Parses something like '1200 rent monthly payment' into an amount, a
    best-guess category match, and whatever text is left over as the note.
    Returns (amount_or_None, category_id_or_None, note). Never guesses wrong
    silently — the caller shows the parsed result for confirmation before
    actually logging anything."""
    text = (text or "").strip()
    amount = None
    match = re.search(r"(\d+(?:\.\d+)?)", text)
    if match:
        amount = float(match.group(1))
        text = (text[:match.start()] + text[match.end():]).strip()

    matched_category_id = None
    matched_len = 0
    remainder = text
    for c in categories:
        name = c["name"].strip()
        if not name:
            continue
        idx = text.lower().find(name.lower())
        if idx != -1 and len(name) > matched_len:
            matched_category_id = c["id"]
            matched_len = len(name)
            remainder = (text[:idx] + text[idx + len(name):]).strip()

    note = re.sub(r"\s+", " ", remainder).strip(" -,.")
    return amount, matched_category_id, note


def guess_category_for_description(description, categories):
    """Best-effort category guess for a bank statement row, based on whether a
    category's name appears in the transaction description. Returns a category
    id, or None if nothing matches — left for the user to assign manually."""
    desc = (description or "").lower()
    for c in categories:
        if c["name"].strip() and c["name"].strip().lower() in desc:
            return c["id"]
    return None


def _guess_column(headers, keywords):
    for h in headers:
        if any(k in h.lower() for k in keywords):
            return h
    return headers[0] if headers else None


def parse_bank_amount(raw):
    """Strips currency symbols/commas/spaces and returns a positive float, since
    bank exports vary between showing debits as negative or as a separate column."""
    cleaned = re.sub(r"[^\d.\-]", "", str(raw or ""))
    try:
        return abs(float(cleaned)) if cleaned not in ("", "-", ".") else None
    except ValueError:
        return None


def parse_bank_date(raw):
    """Tries a handful of common bank statement date formats and normalizes to
    YYYY-MM-DD. Returns None if nothing matches, so the caller can flag the row
    instead of silently importing a wrong date."""
    raw = (raw or "").strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%d %B %Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(raw, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def page_budget(data):
    st.title("Budget")

    months = available_transaction_months(data)
    if "budget_month" not in st.session_state or st.session_state.budget_month not in months:
        st.session_state.budget_month = current_month_str()
    selected_month = st.selectbox(
        "Viewing month", months, key="budget_month",
        format_func=lambda m: date.fromisoformat(m + "-01").strftime("%B %Y"),
    )
    is_current_month = selected_month == current_month_str()

    spent = total_spent(data, selected_month)
    budgeted = total_budgeted(data)
    st.caption(f"Income {fmt(data['income'])} · Budgeted {fmt(budgeted)} · Spent {fmt(spent)}")
    if not is_current_month:
        st.caption("📅 Viewing a past month — budgets shown are today's category budgets applied retroactively, not necessarily what they were back then.")

    st.markdown("#### Monthly income")
    new_income = st.number_input("Monthly income", min_value=0, value=int(data["income"]), step=50, label_visibility="collapsed")
    if new_income != data["income"]:
        data["income"] = new_income
        save_current_data(data)
        st.rerun()

    st.markdown("#### Categories")
    for c in data["categories"]:
        s = spent_by_category(data, c["id"], selected_month)
        pct = min(100, round(s / c["budgeted"] * 100)) if c["budgeted"] > 0 else 0
        over = s > c["budgeted"]
        col1, col2 = st.columns([5, 1])
        with col1:
            color = "#EF4444" if over else "#F1F5F9"
            st.markdown(
                f"<div><b>{c['name']}</b></div>"
                f"<div><span class='ll-mono' style='color:{color};'>{fmt(s)}</span> "
                f"<span class='ll-meta'>/ {fmt(c['budgeted'])}</span></div>",
                unsafe_allow_html=True,
            )
            st.progress(pct / 100)
        with col2:
            if st.button("Delete", key=f"delcat_{c['id']}"):
                data["categories"] = [x for x in data["categories"] if x["id"] != c["id"]]
                data["transactions"] = [t for t in data["transactions"] if t["categoryId"] != c["id"]]
                save_current_data(data)
                st.rerun()
    if not data["categories"]:
        st.caption("No categories yet. Add one below.")

    with st.form("add_category_form", clear_on_submit=True):
        c1, c2, c3 = st.columns([2, 2, 1])
        name = c1.text_input("Category name")
        budget_amt = c2.number_input("Budgeted amount", min_value=0, step=10)
        submitted = c3.form_submit_button("Add category")
        if submitted and name and budget_amt >= 0:
            data["categories"].append({"id": uid(), "name": name, "budgeted": budget_amt})
            save_current_data(data)
            st.rerun()

    st.markdown("#### Recurring transactions")
    st.caption("Set up monthly expenses once (like rent) and they'll auto-log each month without you re-entering them.")
    for r in data["recurring_transactions"]:
        cat = next((c for c in data["categories"] if c["id"] == r["categoryId"]), None)
        cname = cat["name"] if cat else "Uncategorized"
        col1, col2 = st.columns([5, 1])
        col1.markdown(
            f'<div class="ll-row"><div>{cname}<div class="ll-meta">Day {r["day_of_month"]} of each month'
            f'{" · " + r["note"] if r.get("note") else ""}</div></div>'
            f'<div class="ll-mono" style="color:var(--brick);">-{fmt(r["amount"])}</div></div>',
            unsafe_allow_html=True,
        )
        if col2.button("Delete", key=f"delrecur_{r['id']}"):
            data["recurring_transactions"] = [x for x in data["recurring_transactions"] if x["id"] != r["id"]]
            save_current_data(data)
            st.rerun()
    if not data["recurring_transactions"]:
        st.caption("No recurring transactions set up yet.")

    if data["categories"]:
        with st.form("add_recurring_form", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns([2, 1, 1, 2])
            cat_names_r = {c["name"]: c["id"] for c in data["categories"]}
            rcat_choice = c1.selectbox("Category", list(cat_names_r.keys()), key="recur_cat")
            ramount = c2.number_input("Amount", min_value=0, step=1, key="recur_amount")
            rday = c3.number_input("Day of month", min_value=1, max_value=28, value=1, key="recur_day")
            rnote = c4.text_input("Note (optional)", key="recur_note")
            submitted = st.form_submit_button("Add recurring transaction")
            if submitted and ramount > 0:
                data["recurring_transactions"].append({
                    "id": uid(), "categoryId": cat_names_r[rcat_choice],
                    "amount": ramount, "day_of_month": int(rday), "note": rnote,
                })
                save_current_data(data)
                st.rerun()

    st.markdown(f"#### Transactions — {date.fromisoformat(selected_month + '-01').strftime('%B %Y')}")
    month_transactions = sorted(transactions_for_month(data, selected_month), key=lambda t: t["date"], reverse=True)
    tx_page_key = f"tx_page_limit_{selected_month}"
    shown_tx, remaining_tx = paginated(month_transactions, tx_page_key)
    for t in shown_tx:
        cat = next((c for c in data["categories"] if c["id"] == t["categoryId"]), None)
        cname = cat["name"] if cat else "Uncategorized"
        note = f" · {t['note']}" if t.get("note") else ""
        tag = " <span class='ll-tag' style='background:var(--forest);'>🔁 recurring</span>" if t.get("recurringId") else ""
        col1, col2 = st.columns([5, 1])
        col1.markdown(f'<div class="ll-row"><div>{cname}{tag}<div class="ll-meta">{t["date"]}{note}</div></div>'
                      f'<div class="ll-mono" style="color:var(--brick);">-{fmt(t["amount"])}</div></div>',
                      unsafe_allow_html=True)
        if col2.button("Delete", key=f"deltx_{t['id']}"):
            data["transactions"] = [x for x in data["transactions"] if x["id"] != t["id"]]
            save_current_data(data)
            st.rerun()
    render_show_more(tx_page_key, remaining_tx)
    if not month_transactions:
        st.caption("No transactions logged for this month yet.")

    if data["categories"]:
        if st.session_state.pop("clear_quick_add", False):
            st.session_state.quick_add_text = ""

        st.markdown("**Quick add**")
        quick_text = st.text_input(
            "Quick add", key="quick_add_text", label_visibility="collapsed",
            placeholder="e.g. '1200 rent' or '450 groceries lunch with friends'",
        )
        if quick_text.strip():
            qamount, qcat_id, qnote = parse_quick_add(quick_text, data["categories"])
            qcat = next((c for c in data["categories"] if c["id"] == qcat_id), None)
            p1, p2, p3, p4 = st.columns([1, 1.5, 2, 1])
            p1.markdown(f"<div class='ll-meta'>Amount</div><div class='ll-mono'>{fmt(qamount) if qamount is not None else '—'}</div>", unsafe_allow_html=True)
            cat_names_q = {c["name"]: c["id"] for c in data["categories"]}
            default_idx = list(cat_names_q.values()).index(qcat_id) if qcat_id in cat_names_q.values() else 0
            confirm_cat = p2.selectbox("Category", list(cat_names_q.keys()), index=default_idx, key="quick_add_cat", label_visibility="collapsed")
            p3.markdown(f"<div class='ll-meta'>Note</div>{qnote or '—'}", unsafe_allow_html=True)
            if p4.button("Add", key="quick_add_confirm", use_container_width=True):
                if qamount is None or qamount <= 0:
                    st.error("Couldn't find an amount — try including a number, like '1200 rent'.")
                else:
                    data["transactions"].append({
                        "id": uid(), "date": str(date.today()), "categoryId": cat_names_q[confirm_cat],
                        "amount": qamount, "note": qnote,
                    })
                    save_current_data(data)
                    st.session_state.clear_quick_add = True
                    st.rerun()
        st.caption("Type an amount plus a category name and it'll parse automatically — review the guess, then click Add.")

    if data["categories"]:
        with st.form("add_tx_form", clear_on_submit=True):
            c1, c2, c3, c4 = st.columns([1.3, 2, 1, 2])
            tdate = c1.date_input("Date", value=date.today())
            cat_names = {c["name"]: c["id"] for c in data["categories"]}
            cat_choice = c2.selectbox("Category", list(cat_names.keys()))
            amount = c3.number_input("Amount", min_value=0, step=1)
            tnote = c4.text_input("Note (optional)")
            submitted = st.form_submit_button("Add transaction")
            if submitted and amount > 0:
                data["transactions"].append({
                    "id": uid(), "date": str(tdate), "categoryId": cat_names[cat_choice],
                    "amount": amount, "note": tnote,
                })
                save_current_data(data)
                st.rerun()
    else:
        st.caption("Add a category first before logging transactions.")

    if data["categories"]:
        with st.expander("Import transactions from a bank statement (CSV)"):
            st.caption(
                "Works with most bank CSV exports. Bank formats vary, so you'll map which column is which, "
                "preview the parsed rows, and can fix categories or skip rows before anything is actually added."
            )
            bank_file = st.file_uploader("Bank statement CSV", type="csv", key="bank_csv_upload")
            if bank_file is not None:
                try:
                    text = bank_file.getvalue().decode("utf-8-sig")
                    reader = csv.DictReader(io.StringIO(text))
                    headers = reader.fieldnames or []
                    rows = list(reader)
                except Exception:
                    headers, rows = [], []
                    st.error("Couldn't read that file — make sure it's a plain CSV export.")

                if headers and rows:
                    c1, c2, c3 = st.columns(3)
                    date_col = c1.selectbox("Date column", headers, index=headers.index(_guess_column(headers, ["date"])), key="bank_date_col")
                    amount_col = c2.selectbox("Amount column", headers, index=headers.index(_guess_column(headers, ["amount", "debit", "value", "withdrawal"])), key="bank_amount_col")
                    desc_col = c3.selectbox("Description column", headers, index=headers.index(_guess_column(headers, ["description", "narrative", "details", "memo"])), key="bank_desc_col")

                    cat_names_b = {c["name"]: c["id"] for c in data["categories"]}
                    cat_options = ["(skip this row)"] + list(cat_names_b.keys())
                    preview = []
                    for row in rows:
                        parsed_date = parse_bank_date(row.get(date_col, ""))
                        parsed_amount = parse_bank_amount(row.get(amount_col, ""))
                        desc = row.get(desc_col, "")
                        guessed_id = guess_category_for_description(desc, data["categories"])
                        guessed_name = next((c["name"] for c in data["categories"] if c["id"] == guessed_id), "(skip this row)")
                        preview.append({
                            "Import": parsed_date is not None and parsed_amount is not None,
                            "Date": parsed_date or "⚠️ unrecognized",
                            "Amount": parsed_amount if parsed_amount is not None else 0.0,
                            "Description": desc,
                            "Category": guessed_name,
                        })

                    st.caption(f"{len(preview)} row(s) found. Review below — uncheck any row you don't want, and fix categories as needed.")
                    edited = st.data_editor(
                        preview,
                        column_config={
                            "Import": st.column_config.CheckboxColumn("Import?"),
                            "Category": st.column_config.SelectboxColumn("Category", options=cat_options),
                            "Amount": st.column_config.NumberColumn("Amount", format="%.2f"),
                        },
                        disabled=["Date", "Description"],
                        hide_index=True,
                        key="bank_import_editor",
                    )

                    if st.button("Import checked rows"):
                        added = 0
                        for row in edited:
                            if not row["Import"] or row["Category"] == "(skip this row)":
                                continue
                            if row["Date"] == "⚠️ unrecognized" or row["Amount"] <= 0:
                                continue
                            data["transactions"].append({
                                "id": uid(), "date": row["Date"], "categoryId": cat_names_b[row["Category"]],
                                "amount": row["Amount"], "note": row["Description"][:80],
                            })
                            added += 1
                        if added:
                            save_current_data(data)
                            st.success(f"Imported {added} transaction(s).")
                            st.rerun()
                        else:
                            st.warning("Nothing was imported — check that rows are ticked and have a category assigned.")
                elif bank_file is not None:
                    st.warning("That file appears to be empty or not a valid CSV.")


# ---------------------------------------------------------------------------
# Page: Goals
# ---------------------------------------------------------------------------

def page_goals(data):
    st.title("Goals")

    if st.session_state.get("goal_notification"):
        st.success(st.session_state.goal_notification)
        st.balloons()
        st.session_state.goal_notification = None

    total_target = sum(g["target"] for g in data["goals"])
    total_saved = sum(g["saved"] for g in data["goals"])
    st.caption(f"{fmt(total_saved)} saved of {fmt(total_target)} across {len(data['goals'])} goal(s)")
    st.markdown("Track targets like an emergency fund or a big purchase, separately from your investment portfolio.")

    for g in data["goals"]:
        pct = min(100, round(g["saved"] / g["target"] * 100)) if g["target"] > 0 else 0
        done = g["saved"] >= g["target"] and g["target"] > 0
        col1, col2 = st.columns([5, 1])
        with col1:
            label = f"<b>{g['name']}</b>" + (" — reached! 🎉" if done else "")
            color = "#22C55E" if done else "#F1F5F9"
            st.markdown(
                f"<div>{label}</div>"
                f"<div><span class='ll-mono' style='color:{color};'>{fmt(g['saved'])}</span> "
                f"<span class='ll-meta'>/ {fmt(g['target'])}</span></div>",
                unsafe_allow_html=True,
            )
            st.progress(pct / 100)
            add_col, btn_col = st.columns([2, 1])
            add_amt = add_col.number_input("Add funds", min_value=0, step=10, key=f"addgoal_{g['id']}", label_visibility="collapsed")
            if btn_col.button("Add funds", key=f"addgoalbtn_{g['id']}"):
                if add_amt > 0:
                    g["saved"] += add_amt
                    now_done = g["saved"] >= g["target"] and g["target"] > 0
                    if now_done and not g.get("notified"):
                        g["notified"] = True
                        st.session_state.goal_notification = f'🎉 Goal reached! "{g["name"]}" is fully funded at {fmt(g["saved"])}.'
                    elif not now_done:
                        g["notified"] = False
                    save_current_data(data)
                    st.rerun()
        with col2:
            if st.button("Delete", key=f"delgoal_{g['id']}"):
                data["goals"] = [x for x in data["goals"] if x["id"] != g["id"]]
                save_current_data(data)
                st.rerun()
    if not data["goals"]:
        st.caption("No goals yet. Add one below — like an emergency fund or a vacation.")

    with st.form("add_goal_form", clear_on_submit=True):
        c1, c2, c3 = st.columns([2, 2, 1])
        name = c1.text_input("Goal name")
        target = c2.number_input("Target amount", min_value=0, step=50)
        submitted = c3.form_submit_button("Add goal")
        if submitted and name and target > 0:
            data["goals"].append({"id": uid(), "name": name, "target": target, "saved": 0, "notified": False})
            save_current_data(data)
            st.rerun()


# ---------------------------------------------------------------------------
# Page: Risk Profile
# ---------------------------------------------------------------------------

def page_risk_profile(data):
    st.title("Risk Profile")
    profile = data.get("risk_profile")

    if profile and not st.session_state.get("retaking_risk_quiz"):
        level = profile["level"]
        st.markdown(f"### You're **{level}**")
        st.caption(f"Score {profile['score']}/24 · taken {profile.get('date', 'earlier')}")
        st.markdown(RISK_LEVEL_BLURB[level])

        st.markdown("#### An illustrative allocation for this profile")
        alloc = RISK_ALLOCATIONS[level]
        for asset_type, pct in alloc.items():
            if pct <= 0:
                continue
            c1, c2 = st.columns([1, 5])
            c1.markdown(f"<div class='ll-mono'>{pct}%</div>", unsafe_allow_html=True)
            c2.progress(pct / 100, text=TYPE_LABELS.get(asset_type, asset_type))
        st.markdown(
            "<p style='font-style: italic; color: var(--brick); font-size: 0.85rem;'>"
            "An illustrative starting point based on your answers, not personalized financial advice — "
            "I'm not a financial advisor. Your actual allocation should also depend on things this quiz "
            "doesn't capture, like your specific goals and full financial picture."
            "</p>",
            unsafe_allow_html=True,
        )

        st.markdown("#### Where this profile points")
        tabs_for_level = RISK_LEVEL_TABS.get(level, [])
        cols = st.columns(len(tabs_for_level))
        for col, region in zip(cols, tabs_for_level):
            if col.button(f"{region} →", key=f"risk_jump_{region}", use_container_width=True):
                st.session_state.broker_region_tab = region
                st.session_state.broker_search = ""
                st.session_state.pending_nav = "Portfolio"
                st.rerun()

        if st.button("Retake the quiz"):
            st.session_state.retaking_risk_quiz = True
            st.rerun()
        return

    st.caption("6 quick questions to get a general sense of your risk tolerance — used to tailor investment suggestions elsewhere in the app.")
    with st.form("risk_quiz_form"):
        answers = []
        for i, q in enumerate(RISK_QUESTIONS):
            labels = [opt[0] for opt in q["options"]]
            choice = st.radio(q["question"], labels, key=f"riskq_{i}", index=None)
            answers.append((q, choice))
        submitted = st.form_submit_button("See my risk profile", type="primary")
        if submitted:
            if any(choice is None for _, choice in answers):
                st.error("Please answer every question.")
            else:
                points = []
                for q, choice in answers:
                    pts = next(p for label, p in q["options"] if label == choice)
                    points.append(pts)
                score, level = score_risk_answers(points)
                data["risk_profile"] = {"score": score, "level": level, "date": str(date.today())}
                save_current_data(data)
                st.session_state.retaking_risk_quiz = False
                st.rerun()


# ---------------------------------------------------------------------------
# Page: Portfolio
# ---------------------------------------------------------------------------

def page_portfolio(data):
    st.title("Portfolio")
    total = total_portfolio(data)
    st.caption(f"Total value {fmt(total)}")

    # --- Net worth goal ---
    st.markdown("#### Net worth goal")
    new_goal = st.number_input("Net worth goal", min_value=0, value=int(data["net_worth_goal"]), step=1000, label_visibility="collapsed")
    if new_goal != data["net_worth_goal"]:
        data["net_worth_goal"] = new_goal
        save_current_data(data)
        st.rerun()
    goal_pct = min(100, round(total / data["net_worth_goal"] * 100)) if data["net_worth_goal"] > 0 else 0
    st.caption(f"{fmt(total)} of {fmt(data['net_worth_goal'])} · {goal_pct}%")
    st.progress(goal_pct / 100)

    history = data.get("net_worth_history", [])
    if len(history) >= 2 and go is not None:
        fig_nw = go.Figure(go.Scatter(
            x=[h["month"] for h in history], y=[h["value"] for h in history],
            mode="lines+markers", line=dict(color="#A78BFA", width=2), marker=dict(size=6, color="#A78BFA"),
            fill="tozeroy", fillcolor="rgba(167,139,250,0.12)",
        ))
        fig_nw.update_layout(
            height=180, margin=dict(l=10, r=10, t=10, b=10),
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)", font=dict(family="Inter"),
            xaxis=dict(showgrid=False, color="#94A3B8"), yaxis=dict(showgrid=True, gridcolor="#2E3B52", color="#94A3B8"),
        )
        st.plotly_chart(fig_nw, use_container_width=True)
    elif len(history) < 2:
        st.caption("Net worth history will appear here once there's more than one month of data.")

    # --- Allocation ---
    st.markdown("#### Allocation")
    st.caption("Pick a type below to filter holdings.")
    by_type = {}
    for h in data["holdings"]:
        by_type[h["type"]] = by_type.get(h["type"], 0) + h["value"]

    if data["holdings"] and go is not None:
        fig = go.Figure(go.Pie(
            labels=[TYPE_LABELS[t] for t in by_type],
            values=list(by_type.values()),
            hole=0.6,
            marker=dict(colors=[TYPE_COLORS[t] for t in by_type]),
        ))
        fig.update_layout(height=260, margin=dict(l=10, r=10, t=10, b=10),
                           paper_bgcolor="rgba(0,0,0,0)", showlegend=True, font=dict(family="Inter"),
                           annotations=[dict(text=fmt(total), x=0.5, y=0.5, font_size=18, showarrow=False)])
        st.plotly_chart(fig, use_container_width=True)
    elif not data["holdings"]:
        st.caption("Add holdings below to see your allocation.")

    type_options = ["All"] + [TYPE_LABELS[t] for t in by_type]
    if "holdings_type_filter" not in st.session_state:
        st.session_state.holdings_type_filter = "All"
    st.session_state.holdings_type_filter = st.radio(
        "Filter by type", type_options, horizontal=True,
        index=type_options.index(st.session_state.holdings_type_filter) if st.session_state.holdings_type_filter in type_options else 0,
        label_visibility="collapsed",
    )

    # --- Holdings (filtered by type, connected to allocation above) ---
    st.markdown("#### Holdings")
    label_filter = st.session_state.holdings_type_filter
    shown_holdings = data["holdings"]
    if label_filter != "All":
        type_key = next(k for k, v in TYPE_LABELS.items() if v == label_filter)
        shown_holdings = [h for h in data["holdings"] if h["type"] == type_key]

    if "contrib_filter_holding_id" not in st.session_state:
        st.session_state.contrib_filter_holding_id = None

    for h in shown_holdings:
        pct = round(h["value"] / total * 100) if total > 0 else 0
        col1, col2, col3 = st.columns([4, 1.5, 1])
        col1.markdown(f'<div class="ll-row" style="border-bottom:none;"><div>{h["name"]}'
                      f'<div class="ll-meta">{TYPE_LABELS[h["type"]]} · {pct}% of portfolio</div></div>'
                      f'<div class="ll-mono">{fmt(h["value"])}</div></div>', unsafe_allow_html=True)
        if col2.button("View contributions", key=f"viewcontrib_{h['id']}"):
            st.session_state.contrib_filter_holding_id = (
                None if st.session_state.contrib_filter_holding_id == h["id"] else h["id"]
            )
            st.rerun()
        if col3.button("Delete", key=f"delhold_{h['id']}"):
            data["holdings"] = [x for x in data["holdings"] if x["id"] != h["id"]]
            for c in data["contributions"]:
                if c.get("holdingId") == h["id"]:
                    c["holdingId"] = None
            save_current_data(data)
            st.rerun()
    if not shown_holdings:
        st.caption("No holdings match this filter." if data["holdings"] else "No holdings yet. Add one below.")

    default_name = st.session_state.pop("prefill_holding_name", "")
    with st.form("add_holding_form", clear_on_submit=True):
        c1, c2, c3 = st.columns([2, 1.5, 1.5])
        name = c1.text_input("Holding name", value=default_name)
        htype = c2.selectbox("Type", list(TYPE_LABELS.keys()), format_func=lambda k: TYPE_LABELS[k])
        value = c3.number_input("Current value", min_value=0, step=10)
        submitted = st.form_submit_button("Add holding")
        if submitted and name and value >= 0:
            data["holdings"].append({"id": uid(), "name": name, "type": htype, "value": value})
            save_current_data(data)
            st.rerun()

    # --- Contributions (connected to the holding selected above) ---
    st.markdown("#### Contributions")
    filter_id = st.session_state.contrib_filter_holding_id
    shown_contribs = data["contributions"]
    if filter_id:
        fh = next((h for h in data["holdings"] if h["id"] == filter_id), None)
        if fh:
            st.caption(f"Showing contributions for **{fh['name']}** — click 'View contributions' again to clear.")
            shown_contribs = [c for c in data["contributions"] if c.get("holdingId") == filter_id]

    shown_contribs = sorted(shown_contribs, key=lambda c: c["date"], reverse=True)
    contrib_page_key = f"contrib_page_limit_{filter_id or 'all'}"
    shown_page, remaining_contribs = paginated(shown_contribs, contrib_page_key)
    for c in shown_page:
        h = next((h for h in data["holdings"] if h["id"] == c.get("holdingId")), None)
        name = h["name"] if h else "Unallocated"
        note = f" · {c['note']}" if c.get("note") else ""
        col1, col2 = st.columns([5, 1])
        col1.markdown(f'<div class="ll-row"><div>{name}<div class="ll-meta">{c["date"]}{note}</div></div>'
                      f'<div class="ll-mono" style="color:var(--forest);">+{fmt(c["amount"])}</div></div>',
                      unsafe_allow_html=True)
        if col2.button("Delete", key=f"delcontrib_{c['id']}"):
            if c.get("holdingId"):
                hh = next((h for h in data["holdings"] if h["id"] == c["holdingId"]), None)
                if hh:
                    hh["value"] = max(0, hh["value"] - c["amount"])
            data["contributions"] = [x for x in data["contributions"] if x["id"] != c["id"]]
            save_current_data(data)
            st.rerun()
    render_show_more(contrib_page_key, remaining_contribs)
    if not shown_contribs:
        st.caption("No contributions logged yet.")

    with st.form("add_contribution_form", clear_on_submit=True):
        c1, c2, c3, c4 = st.columns([1.3, 1, 2, 2])
        cdate = c1.date_input("Date", value=date.today(), key="cdate")
        camount = c2.number_input("Amount", min_value=0, step=10, key="camount")
        holding_names = {"Unallocated": None}
        holding_names.update({h["name"]: h["id"] for h in data["holdings"]})
        choice = c3.selectbox("Holding", list(holding_names.keys()))
        cnote = c4.text_input("Note (optional)", key="cnote")
        submitted = st.form_submit_button("Log contribution")
        if submitted and camount > 0:
            hid = holding_names[choice]
            data["contributions"].append({"id": uid(), "date": str(cdate), "amount": camount, "note": cnote, "holdingId": hid})
            if hid:
                hh = next(h for h in data["holdings"] if h["id"] == hid)
                hh["value"] += camount
            save_current_data(data)
            st.rerun()

    saved = saved_this_month(data, current_month_str())
    if saved > 0:
        if st.button(f"Log this month's available {fmt(saved)} as a contribution"):
            data["contributions"].append({"id": uid(), "date": str(date.today()), "amount": saved, "note": "Available income invested", "holdingId": None})
            save_current_data(data)
            st.rerun()

    # --- Where to open an account ---
    st.markdown("#### Where to open an account")
    st.caption("Ranked by the size of each region's retail investor base. This isn't a recommendation — compare fees and features before choosing.")
    if not st.session_state.get("broker_directory_flagged"):
        data.setdefault("usage", {})["opened_broker_directory"] = True
        st.session_state.broker_directory_flagged = True
        save_current_data(data)
    search = st.text_input("Search companies...", key="broker_search", label_visibility="collapsed", placeholder="Search companies...")

    if search.strip():
        q = search.strip().lower()
        matches = [b for b in BROKERS if q in b["name"].lower()]
        st.caption(f"{len(matches)} result(s) across all categories for \"{search}\"")
        for b in matches:
            col1, col2, col3 = st.columns([4, 1.3, 1.3])
            col1.markdown(f'<div class="ll-row" style="border-bottom:none;"><div>{b["name"]}'
                          f'<div class="ll-meta">{b["note"]}</div>'
                          f'<span class="ll-tag">{b["region"]}</span></div></div>', unsafe_allow_html=True)
            col2.link_button("Visit site", b["url"])
            if col3.button("Use as holding", key=f"usehold_{b['name']}"):
                st.session_state.prefill_holding_name = b["name"]
                st.rerun()
        if not matches:
            st.caption(f'No companies match "{search}".')
    else:
        if "broker_region_tab" not in st.session_state:
            st.session_state.broker_region_tab = BROKER_REGIONS[0]
        tab_cols = st.columns(len(BROKER_REGIONS))
        for col, region in zip(tab_cols, BROKER_REGIONS):
            is_active = st.session_state.broker_region_tab == region
            if col.button(region, key=f"tabbtn_{region}", type="primary" if is_active else "secondary", use_container_width=True):
                st.session_state.broker_region_tab = region
                st.rerun()

        active_region = st.session_state.broker_region_tab
        for b in [x for x in BROKERS if x["region"] == active_region]:
            col1, col2, col3 = st.columns([4, 1.3, 1.3])
            col1.markdown(f'<div class="ll-row" style="border-bottom:none;"><div>{b["name"]}'
                          f'<div class="ll-meta">{b["note"]}</div></div></div>', unsafe_allow_html=True)
            col2.link_button("Visit site", b["url"])
            if col3.button("Use as holding", key=f"usehold2_{b['name']}"):
                st.session_state.prefill_holding_name = b["name"]
                st.rerun()


# ---------------------------------------------------------------------------
# Admin usage overview
# ---------------------------------------------------------------------------

def render_admin_page():
    st.title("Admin — Usage Overview")
    st.caption("Only visible to the configured admin account. Aggregated across every Firestore user document.")

    db = get_firestore_client()
    if db is None:
        st.warning("Firestore isn't connected, so there's nothing to aggregate across users yet.")
        return

    try:
        docs = list(db.collection(FIRESTORE_COLLECTION).stream())
    except Exception as e:
        st.error(f"Couldn't read the Firestore collection: {e}")
        return

    user_docs = [d.to_dict() for d in docs if d.to_dict() and "income" in d.to_dict()]
    total_users = len(user_docs)

    if total_users == 0:
        st.info("No user data found yet.")
        return

    total_sessions = sum(u.get("usage", {}).get("session_count", 0) for u in user_docs)
    opened_directory_count = sum(1 for u in user_docs if u.get("usage", {}).get("opened_broker_directory"))
    users_with_goals = sum(1 for u in user_docs if len(u.get("goals", [])) > 0)
    users_with_holdings = sum(1 for u in user_docs if len(u.get("holdings", [])) > 0)
    total_transactions = sum(len(u.get("transactions", [])) for u in user_docs)
    avg_transactions = total_transactions / total_users if total_users else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total users", total_users)
    c2.metric("Total sessions", total_sessions)
    c3.metric("Avg transactions / user", f"{avg_transactions:.1f}")
    c4.metric("Opened broker directory", f"{opened_directory_count}/{total_users}")

    c5, c6 = st.columns(2)
    c5.metric("Users with a goal set", f"{users_with_goals}/{total_users}")
    c6.metric("Users with holdings", f"{users_with_holdings}/{total_users}")

    st.markdown("#### Per-user summary")
    rows = []
    for i, u in enumerate(user_docs):
        usage = u.get("usage", {})
        rows.append({
            "User #": i + 1,
            "Sessions": usage.get("session_count", 0),
            "Last active": usage.get("last_active") or "—",
            "Transactions": len(u.get("transactions", [])),
            "Goals": len(u.get("goals", [])),
            "Portfolio value": total_portfolio(u),
            "Opened directory": "Yes" if usage.get("opened_broker_directory") else "No",
        })
    st.dataframe(rows, hide_index=True, use_container_width=True)
    st.caption("User # is just a row number here, not tied to any identity — emails aren't stored in Firestore, only in Firebase Authentication.")

    st.markdown("#### Ratings & feedback")
    try:
        feedback_docs = list(db.collection(FEEDBACK_COLLECTION).stream())
        feedback_entries = [d.to_dict() for d in feedback_docs if d.to_dict()]
    except Exception as e:
        feedback_entries = []
        st.error(f"Couldn't read feedback: {e}")

    if not feedback_entries:
        st.caption("No ratings submitted yet.")
    else:
        ratings = [f["rating"] for f in feedback_entries if f.get("rating") is not None]
        avg_rating = sum(ratings) / len(ratings) if ratings else 0
        fc1, fc2 = st.columns(2)
        fc1.metric("Average rating", f"{avg_rating:.1f} / 5 ⭐")
        fc2.metric("Total ratings", len(feedback_entries))

        feedback_rows = sorted(feedback_entries, key=lambda f: f.get("date", ""), reverse=True)
        st.dataframe(
            [{"Date": f.get("date", "—"), "Rating": f.get("rating", "—"), "Comment": f.get("comment", "") or "—"} for f in feedback_rows],
            hide_index=True, use_container_width=True,
        )


# ---------------------------------------------------------------------------
# Authentication gate
# ---------------------------------------------------------------------------

def render_auth_gate():
    """Shows a login/signup form and halts the rest of the app until the
    visitor is authenticated. Only used when Firebase + a Web API key are
    configured; otherwise the app falls back to the anonymous URL-code mode."""
    st.title("Ledger Line")
    st.caption("Budget in, invest forward")
    st.markdown("#### Log in or create an account to continue")

    tab_login, tab_signup = st.tabs(["Log in", "Sign up"])

    with tab_login:
        with st.form("login_form"):
            email = st.text_input("Email", key="login_email")
            password = st.text_input("Password", type="password", key="login_password")
            submitted = st.form_submit_button("Log in", type="primary")
            if submitted:
                if not EMAIL_RE.match(email or ""):
                    st.error("Please enter a valid email address.")
                elif not password:
                    st.error("Please enter your password.")
                else:
                    with st.spinner("Logging in..."):
                        ok, result = firebase_sign_in(email, password)
                    if ok:
                        st.query_params["token"] = result["idToken"]
                        st.query_params["rtoken"] = result["refreshToken"]
                        st.session_state.user_email = result.get("email", email)
                        st.session_state.anon_id = result["localId"]
                        st.session_state.authenticated = True
                        st.rerun()
                    else:
                        st.error(result)

        with st.expander("Forgot your password?"):
            with st.form("forgot_password_form"):
                reset_email = st.text_input("Email", key="reset_email")
                reset_submitted = st.form_submit_button("Send reset link")
                if reset_submitted:
                    if not EMAIL_RE.match(reset_email or ""):
                        st.error("Please enter a valid email address.")
                    else:
                        with st.spinner("Sending..."):
                            ok, result = firebase_send_password_reset(reset_email)
                        if ok:
                            st.success(f"If an account exists for {reset_email}, a reset link has been sent — check your inbox (and spam folder).")
                        else:
                            st.error(result)

    with tab_signup:
        with st.form("signup_form"):
            new_email = st.text_input("Email", key="signup_email")
            new_password = st.text_input("Password", type="password", key="signup_password",
                                          help="At least 6 characters.")
            confirm_password = st.text_input("Confirm password", type="password", key="signup_confirm")
            submitted = st.form_submit_button("Create account", type="primary")
            if submitted:
                if not EMAIL_RE.match(new_email or ""):
                    st.error("Please enter a valid email address.")
                elif len(new_password or "") < 6:
                    st.error("Password should be at least 6 characters.")
                elif new_password != confirm_password:
                    st.error("Passwords don't match.")
                else:
                    with st.spinner("Creating your account..."):
                        ok, result = firebase_sign_up(new_email, new_password)
                    if ok:
                        st.query_params["token"] = result["idToken"]
                        st.query_params["rtoken"] = result["refreshToken"]
                        st.session_state.user_email = result.get("email", new_email)
                        st.session_state.anon_id = result["localId"]
                        st.session_state.authenticated = True
                        st.success("Account created!")
                        st.rerun()
                    else:
                        st.error(result)

    st.caption("Your data is private to your account and stored in Firestore, isolated from every other user.")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    st.set_page_config(page_title="Ledger Line", layout="wide")
    st.markdown(PALETTE_CSS, unsafe_allow_html=True)

    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False

    if auth_available():
        if not st.session_state.authenticated:
            token = st.query_params.get("token")
            uid = verify_firebase_token(token) if token else None

            if not uid:
                # The ID token is missing/expired — try silently refreshing it
                # with the longer-lived refresh token before giving up and
                # forcing the visitor to log in again from scratch.
                rtoken = st.query_params.get("rtoken")
                if rtoken:
                    ok, result = firebase_refresh_token(rtoken)
                    if ok:
                        new_id_token = result["access_token"]  # refresh endpoint uses different field names than sign-in
                        new_refresh_token = result["refresh_token"]
                        refreshed_uid = verify_firebase_token(new_id_token)
                        if refreshed_uid:
                            st.query_params["token"] = new_id_token
                            st.query_params["rtoken"] = new_refresh_token
                            uid = refreshed_uid

            if uid:
                st.session_state.authenticated = True
                st.session_state.anon_id = uid
            else:
                if token:
                    st.query_params.clear()
                    st.info("Your session expired — please log in again.")
                render_auth_gate()
                st.stop()
    else:
        if "anon_id" not in st.session_state:
            anon_id, is_new_visitor = get_or_create_anon_id()
            st.session_state.anon_id = anon_id
            st.session_state.is_new_visitor = is_new_visitor

    if "data" not in st.session_state:
        st.session_state.data = load_current_data()
    data = st.session_state.data
    data.setdefault("recurring_transactions", [])
    data.setdefault("goals", [])
    data.setdefault("net_worth_goal", 0)
    data.setdefault("net_worth_history", [])
    data.setdefault("usage", {"session_count": 0, "last_active": None, "opened_broker_directory": False})
    data.setdefault("feedback_history", [])
    data.setdefault("feedback_prompt_count", 0)

    needs_save = False
    if generate_recurring_transactions(data):
        needs_save = True
    if record_net_worth_snapshot(data):
        needs_save = True
    if track_session_usage(data):
        needs_save = True

    should_auto_prompt = (
        not data["feedback_history"]
        and not st.session_state.get("rating_popup_shown_this_session")
        and data.get("usage", {}).get("session_count", 0) >= 3
        and data.get("feedback_prompt_count", 0) < 3
    )
    if should_auto_prompt:
        st.session_state.rating_popup_shown_this_session = True
        st.session_state.show_rating_dialog = True
        data["feedback_prompt_count"] = data.get("feedback_prompt_count", 0) + 1
        needs_save = True

    if needs_save:
        save_current_data(data)

    if st.session_state.get("is_new_visitor"):
        st.info(
            "🔑 **This page is now uniquely yours.** No login needed — your data is tied to this exact "
            "web address. **Bookmark this page** (or save the link) to come back to the same data later. "
            "Don't share this link publicly, since anyone with it can see and edit this data. "
            "Use made-up numbers if you're just trying the app out."
        )
        st.session_state.is_new_visitor = False

    st.sidebar.markdown("### Ledger Line")
    st.sidebar.caption("Budget in, invest forward")
    if using_firestore():
        st.sidebar.caption("🟢 Synced to Firebase")
    else:
        st.sidebar.caption("⚪ Saving locally (Firebase not configured)")

    if not auth_available():
        with st.sidebar.expander("Why isn't login showing up?"):
            for label, ok in auth_diagnostics():
                st.markdown(f"{'✅' if ok else '❌'} {label}")

    if st.session_state.authenticated:
        st.sidebar.caption(f"👤 {st.session_state.get('user_email', 'Logged in')}")
        if st.sidebar.button("Log out"):
            st.query_params.clear()
            for key in ["authenticated", "anon_id", "user_email", "data"]:
                st.session_state.pop(key, None)
            st.rerun()
    else:
        st.sidebar.caption(f"🔑 Access code: `{st.session_state.anon_id}`")
        if st.sidebar.button("Start a new anonymous session"):
            st.query_params.clear()
            for key in ["anon_id", "is_new_visitor", "data"]:
                st.session_state.pop(key, None)
            st.rerun()

    if "pending_nav" in st.session_state:
        st.session_state.nav_radio = st.session_state.pop("pending_nav")
    if "nav_radio" not in st.session_state:
        st.session_state.nav_radio = "Overview"
    nav_options = ["Overview", "Budget", "Goals", "Risk Profile", "Portfolio"]
    is_admin = st.session_state.get("authenticated") and st.session_state.get("user_email") and st.session_state.get("user_email") == get_admin_email()
    if is_admin:
        nav_options.append("Admin")
    page = st.sidebar.radio("Navigate", nav_options, key="nav_radio", label_visibility="collapsed")

    st.sidebar.markdown("---")
    st.sidebar.markdown("**Export data**")
    export_format = st.sidebar.selectbox(
        "Export format", ["CSV", "JSON", "Excel (.xlsx)"], label_visibility="collapsed", key="export_format",
    )
    if export_format == "CSV":
        st.sidebar.download_button("Download", data=data_to_csv(data),
                                    file_name=f"ledger-line-export-{date.today()}.csv", mime="text/csv")
    elif export_format == "JSON":
        st.sidebar.download_button("Download", data=json.dumps(data, indent=2),
                                    file_name=f"ledger-line-export-{date.today()}.json", mime="application/json")
    else:
        if openpyxl is not None:
            st.sidebar.download_button(
                "Download", data=data_to_xlsx(data),
                file_name=f"ledger-line-export-{date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.sidebar.caption("Install openpyxl (`pip install openpyxl`) to export Excel files.")

    st.sidebar.markdown("**Import data**")
    st.sidebar.caption("Accepts a .csv, .json, or .xlsx export from this app.")
    uploaded = st.sidebar.file_uploader("Import data", type=["csv", "json", "xlsx"], label_visibility="collapsed")
    if uploaded is not None:
        ext = uploaded.name.rsplit(".", 1)[-1].lower()
        try:
            if ext == "csv":
                parsed = csv_to_data(uploaded.getvalue().decode("utf-8"))
            elif ext == "json":
                parsed = json_to_data(uploaded.getvalue().decode("utf-8"))
            elif ext == "xlsx":
                if openpyxl is None:
                    raise RuntimeError("openpyxl not installed")
                parsed = xlsx_to_data(uploaded.getvalue())
            else:
                parsed = None

            has_content = parsed and (parsed["categories"] or parsed["transactions"] or parsed["holdings"] or parsed["income"])
            if has_content:
                st.session_state.data = parsed
                save_current_data(parsed)
                st.sidebar.success("Data imported.")
                st.rerun()
            else:
                st.sidebar.error("That file doesn't look like a Ledger Line export.")
        except Exception:
            st.sidebar.error("Couldn't read that file. Make sure it's a valid Ledger Line export in .csv, .json, or .xlsx format.")

    if st.sidebar.button("⭐ Rate this app"):
        st.session_state.show_rating_dialog = True

    if st.session_state.get("show_rating_dialog"):
        st.session_state.show_rating_dialog = False  # consume immediately, same reasoning as the reset dialog
        rating_dialog(data)

    if st.sidebar.button("Reset data"):
        st.session_state.show_reset_dialog = True

    if st.session_state.get("show_reset_dialog"):
        st.session_state.show_reset_dialog = False  # consume immediately: dialog stays open on its own until an explicit choice is made
        reset_confirmation_dialog()

    if page == "Overview":
        page_overview(data)
    elif page == "Budget":
        page_budget(data)
    elif page == "Goals":
        page_goals(data)
    elif page == "Risk Profile":
        page_risk_profile(data)
    elif page == "Admin":
        render_admin_page()
    else:
        page_portfolio(data)


if __name__ == "__main__":
    main()
